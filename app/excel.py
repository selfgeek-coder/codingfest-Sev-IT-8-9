from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config import Settings

def export_orders_to_excel(orders, filepath="orders.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = [
        "Дата",
        "ФИ",
        "Telegram username",
        "Статус заказа",
        "Материал",
        "Цвет",
        "Цена за шт.",
        "Итого",
        "Товар",
        "Кол-во",
    ]

    ws.append(headers)

    for order in orders:
        user = order.user  

        row = [
            order.created_at.strftime("%d.%m.%Y %H:%M"),
            order.full_name,
            f"@{user.username}" if user and user.username else "нет",
            Settings.human_status.get(order.status, order.status.value),
            order.material,
            order.color,
            order.unit_price_rub,
            order.price_rub,
            order.name,
            order.quantity,
        ]

        ws.append(row)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(filepath)
    return filepath
